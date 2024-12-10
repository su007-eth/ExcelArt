from PIL import Image
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import numpy as np
import os

class ExcelArt:
    def __init__(self, image_path, max_size=150):
        """
        初始化ExcelArt
        :param image_path: 图片路径
        :param max_size: 处理后图片的最大尺寸（像素数）
        """
        self.image_path = image_path
        self.max_size = max_size
        self.image = None
        self.pixel_data = None

    def gcd(self, a, b):
        """计算两个数的最大公约数"""
        while b:
            a, b = b, a % b
        return a

    def detect_pixel_art(self, tolerance=45):
        """
        检测图片是否为像素图并返回原始尺寸
        :param tolerance: 容许误差，用于判断像素块重复
        :return: 如果是像素图返回(原始宽度, 原始高度)，否则返回None, None
        """
        if self.image is None:
            return None, None

        width, height = self.image.size
        print(f"原始图片尺寸: {width}x{height}")

        # 对大图先进行预处理，缩小到合理尺寸以加快处理速度
        MAX_TEST_SIZE = 1000
        if width > MAX_TEST_SIZE or height > MAX_TEST_SIZE:
            scale = MAX_TEST_SIZE / max(width, height)
            test_width = int(width * scale)
            test_height = int(height * scale)
            test_image = self.image.resize((test_width, test_height), Image.Resampling.LANCZOS)
            print(f"预处理：将图片缩放至 {test_width}x{test_height} 以加快处理速度")
        else:
            test_image = self.image
            test_width, test_height = width, height

        # 快速预检查：如果图片是非像素图，大多数像素都会有细微差异
        sample_size = min(100, test_width, test_height)  # 取样大小
        sample = test_image.resize((sample_size, sample_size), Image.Resampling.LANCZOS)
        sample_restored = sample.resize((sample_size * 2, sample_size * 2), Image.Resampling.NEAREST).resize((sample_size, sample_size), Image.Resampling.LANCZOS)
        
        diff = np.subtract(np.asarray(sample), np.asarray(sample_restored))
        quick_test_diff = np.abs(diff).mean()
        
        # 如果快速测试的差异值很大，可能不是像素图
        if quick_test_diff > tolerance * 1.5:
            print(f"\n快速预检查：差异值 {quick_test_diff:.2f} 远大于容差 {tolerance:.2f}，可能不是像素图")
            print("建议：使用较大的像素尺寸（96x96或128x128）以保持更多细节")
            # 直接返回较大的尺寸
            size = 128 if width > 1000 or height > 1000 else 96
            return size, size

        # 按尺寸从小到大排序，包括更大的尺寸选项
        sizes_with_priority = [
            (8, 1.0),
            (16, 1.0),
            (24, 1.0),
            (28, 1.0),
            (32, 1.0),
            (48, 1.0),
            (64, 1.0),
            (96, 1.0),
            (128, 1.0)
        ]
        
        # 首先测试原始尺寸是否为像素图
        original_is_pixel = False
        
        # 只对小图进行原始尺寸测试
        if width <= 1000 and height <= 1000:
            # 放大2倍，然后缩小回原尺寸，看是否一致
            enlarged = self.image.resize((width * 2, height * 2), Image.Resampling.NEAREST)
            restored = enlarged.resize((width, height), Image.Resampling.NEAREST)
            # 使用numpy的优化方法计算差异
            diff = np.subtract(np.asarray(self.image), np.asarray(restored))
            mean_diff = np.abs(diff).mean()
            
            print(f"测试原始尺寸 {width}x{height}: 差异: {mean_diff:.2f}, 容差: {tolerance:.2f}")
            
            if mean_diff == 0:
                original_is_pixel = True
                original_result = (width, height)
                print(f"发现完美像素图！尺寸: {width}x{height}")
                # 注意：不立即返回，继续测试是否有更小的等效尺寸

        # 使用numpy的广播功能优化差异计算
        test_array = np.asarray(test_image)
        
        # 测试其他标准尺寸
        consecutive_bad_results = 0  # 连续不好的结果计数
        perfect_matches = []  # 存储所有完美匹配（差异值为0）
        good_matches = []    # 存储所有好的匹配（差异值在容差内）

        for size, tolerance_mult in sizes_with_priority:
            # 如果尺寸大于原始图片的一半，跳过
            if size > width // 2 or size > height // 2:
                continue
                
            # 计算最接近的合适尺寸
            target_width = (test_width // size) * size
            target_height = (test_height // size) * size
            
            if target_width > 0 and target_height > 0:
                # 如果需要，先调整到合适的尺寸
                if target_width != test_width or target_height != test_height:
                    current_test = test_image.resize((target_width, target_height), Image.Resampling.LANCZOS)
                    current_array = np.asarray(current_test)
                else:
                    current_array = test_array
                
                test_scale = target_width // size
                small = Image.fromarray(current_array).resize((size, size), Image.Resampling.NEAREST)
                restored = small.resize((target_width, target_height), Image.Resampling.NEAREST)
                
                # 使用numpy的优化方法计算差异
                diff = np.subtract(current_array, np.asarray(restored))
                mean_diff = np.abs(diff).mean()
                
                # 根据预处理的缩放比例调整差异值
                if width > MAX_TEST_SIZE or height > MAX_TEST_SIZE:
                    mean_diff = mean_diff * (1 + scale)  # 补偿缩放带来的误差
                
                print(f"测试 {size}x{size}: 目标尺寸 {target_width}x{target_height}, "
                      f"缩放因子 {test_scale}, 差异: {mean_diff:.2f}, "
                      f"容差: {tolerance:.2f}")
                
                # 记录匹配结果
                if mean_diff == 0:
                    perfect_matches.append((size, size))
                elif mean_diff <= tolerance:
                    good_matches.append((size, size, mean_diff))

                # 如果连续多个结果都很差，可能不是像素图
                if mean_diff > tolerance:
                    consecutive_bad_results += 1
                    if consecutive_bad_results >= 3:  # 连续3个结果都不好
                        print("\n连续多个测试结果较差，可能不是像素图")
                        print("建议：使用较大的像素尺寸以保持更多细节")
                        # 返回较大的尺寸
                        size = 128 if width > 1000 or height > 1000 else 96
                        return size, size
                else:
                    consecutive_bad_results = 0

        # 选择最佳结果
        if perfect_matches:
            # 如果有完美匹配，选择最小的一个
            best_match = min(perfect_matches, key=lambda x: x[0])
            if original_is_pixel:
                # 如果原始尺寸也是完美匹配，比较大小
                if best_match[0] < original_result[0]:
                    print(f"\n发现更小的等效像素尺寸！选择 {best_match[0]}x{best_match[1]}")
                    return best_match
                else:
                    print(f"\n选择原始尺寸 {original_result[0]}x{original_result[1]} 作为最佳匹配")
                    return original_result
            print(f"\n选择 {best_match[0]}x{best_match[1]} 作为最佳匹配（完美像素图）")
            return best_match
        elif original_is_pixel:
            print(f"\n选择原始尺寸 {original_result[0]}x{original_result[1]} 作为最佳匹配")
            return original_result
        elif good_matches:
            # 如果有好的匹配，先找出最小差异值
            best_match = min(good_matches, key=lambda x: x[2])
            # 找出所有相似的匹配（差异值相差不超过5%）
            similar_matches = [m for m in good_matches if abs(m[2] - best_match[2]) <= best_match[2] * 0.05]
            
            if len(similar_matches) > 1:
                # 如果有多个相似的匹配，选择最小的尺寸
                best_match = min(similar_matches, key=lambda x: x[0])
                print(f"\n选择 {best_match[0]}x{best_match[1]} 作为最佳匹配（差异值: {best_match[2]:.2f}）")
                print("注意：发现多个相似的匹配结果（差异值相差不超过5%），选择了最小的尺寸")
            else:
                print(f"\n选择 {best_match[0]}x{best_match[1]} 作为最佳匹配（差异值: {best_match[2]:.2f}）")
            
            return best_match[0], best_match[1]
        else:
            # 如果没有任何好的匹配，返回较大的尺寸
            size = 128 if width > 1000 or height > 1000 else 96
            print(f"\n未找到好的匹配，使用 {size}x{size} 作为默认尺寸")
            return size, size

    def load_and_resize_image(self):
        """加载并调整图片大小"""
        # 打开图片
        self.image = Image.open(self.image_path)
        
        # 转换为RGB模式
        if self.image.mode != 'RGB':
            self.image = self.image.convert('RGB')

        # 检测是否为像素图
        original_width, original_height = self.detect_pixel_art()
        
        if original_width and original_height:
            # 如果是像素图，直接还原到原始大小
            self.image = self.image.resize((original_width, original_height), Image.Resampling.NEAREST)
        else:
            # 如果不是像素图，使用两步法处理以获得更好的质量
            width, height = self.image.size
            
            # 步骤1：先用高质量的LANCZOS算法缩小到目标尺寸的2倍
            target_size = original_width  # 这是从detect_pixel_art返回的大小（96或128）
            aspect_ratio = width / height
            
            if aspect_ratio > 1:  # 宽图
                intermediate_width = target_size * 2
                intermediate_height = int(intermediate_width / aspect_ratio)
            else:  # 高图
                intermediate_height = target_size * 2
                intermediate_width = int(intermediate_height * aspect_ratio)
            
            # 使用LANCZOS算法进行初步缩放
            self.image = self.image.resize((intermediate_width, intermediate_height), 
                                         Image.Resampling.LANCZOS)
            
            # 步骤2：再用NEAREST算法缩小到最终尺寸，这样可以得到更清晰的像素边界
            if aspect_ratio > 1:  # 宽图
                final_width = target_size
                final_height = int(final_width / aspect_ratio)
            else:  # 高图
                final_height = target_size
                final_width = int(final_height * aspect_ratio)
            
            # 确保尺寸至少为1
            final_width = max(1, final_width)
            final_height = max(1, final_height)
            
            # 使用NEAREST算法进行最终缩放
            self.image = self.image.resize((final_width, final_height), 
                                         Image.Resampling.NEAREST)
            
            print(f"非像素图：调整为 {final_width}x{final_height} 以保持最佳显示效果")
        
        # 更新像素数据
        self.pixel_data = np.array(self.image)

    def create_excel(self, output_path):
        """创建Excel文件并填充像素数据"""
        if self.pixel_data is None:
            raise ValueError("请先加载图片")

        # 创建工作簿和工作表
        wb = openpyxl.Workbook()
        ws = wb.active

        # 根据图片大小动态调整单元格大小
        total_pixels = self.pixel_data.shape[0] * self.pixel_data.shape[1]
        if total_pixels > 10000:  # 大图片
            PIXEL_SIZE = 12  # 小单元格
        elif total_pixels > 5000:  # 中等图片
            PIXEL_SIZE = 15  # 中等单元格
        else:  # 小图片
            PIXEL_SIZE = 20  # 大单元格

        # Excel中列宽单位约为8像素，行高单位为1.33像素
        COLUMN_WIDTH = PIXEL_SIZE / 8  # 转换为Excel列宽单位
        ROW_HEIGHT = PIXEL_SIZE * 0.75  # 转换为Excel行高单位（1像素 ≈ 0.75点）

        # 获取工作表的总列数和行数
        total_cols = max(self.pixel_data.shape[1] + 2, 26)  # 至少26列，确保显示A-Z
        total_rows = self.pixel_data.shape[0] + 2  # 额外加2行作为边距

        # 统一设置所有列宽
        for col in range(1, total_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTH

        # 统一设置所有行高
        for row in range(1, total_rows + 1):
            ws.row_dimensions[row].height = ROW_HEIGHT

        # 填充像素数据（从第2行第2列开始，留出边距）
        for i in range(self.pixel_data.shape[0]):
            for j in range(self.pixel_data.shape[1]):
                r, g, b = self.pixel_data[i, j]
                hex_color = f"{r:02x}{g:02x}{b:02x}"
                cell = ws.cell(row=i+2, column=j+2)
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

        # 保存Excel文件
        wb.save(output_path)

def get_default_excel_path(image_path):
    """
    根据输入图片路径生成默认的Excel输出路径
    :param image_path: 输入图片路径
    :return: 默认的Excel输出路径
    """
    # 获取图片所在目录和文件名（不含扩展名）
    directory = os.path.dirname(image_path)
    base_name = os.path.splitext(os.path.basename(image_path))[0]
    # 生成Excel文件路径
    return os.path.join(directory, f"{base_name}.xlsx")

def convert_image_to_excel(image_path, output_path=None, max_size=150):
    """
    将图片转换为Excel艺术图
    :param image_path: 输入图片路径
    :param output_path: 输出Excel文件路径，如果为None则使用默认路径
    :param max_size: 最大像素尺寸
    """
    if output_path is None:
        output_path = get_default_excel_path(image_path)
    
    converter = ExcelArt(image_path, max_size)
    converter.load_and_resize_image()
    converter.create_excel(output_path)

if __name__ == "__main__":
    # 使用示例
    import sys
    if len(sys.argv) < 2 or len(sys.argv) > 3:
        print("使用方法: python excel_art.py 输入图片路径 [输出Excel路径]")
        print("注意：如果不指定输出路径，将在与输入图片相同的目录下创建同名的Excel文件")
        sys.exit(1)
    
    input_image = sys.argv[1]
    output_excel = sys.argv[2] if len(sys.argv) == 3 else None
    convert_image_to_excel(input_image, output_excel)
